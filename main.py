import os
import json
import pandas as pd
import numpy as np
import logging
from time import time
from src.measurements.vsg import VSG
from src.measurements.vsa import VSA
from src.measurements.power_meter import PowerMeter
from src.measurements.power_servo import PowerServo
from openpyxl.styles import Font, PatternFill

# Setup logging
logger = logging.getLogger(__name__)
base_path = os.path.dirname(__file__)
log_dir = os.path.join(base_path, 'logs')
os.makedirs(log_dir, exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(os.path.join(log_dir, 'sweep_measurement.log')),
        logging.StreamHandler()
    ]
)

def run_sweep():
    json_path = os.path.join(base_path, 'test_inputs.json')
    cal_path = os.path.join(base_path, 'combined_cal_data.xlsx')
    output_path = os.path.join(base_path, 'sweep_measurements.xlsx')

    # Check if input files exist
    logger.info(f"Checking for test inputs file: {os.path.abspath(json_path)}")
    if not os.path.exists(json_path):
        logger.error(f"Test inputs file not found: {os.path.abspath(json_path)}")
        return
    logger.info(f"Checking for calibration data file: {os.path.abspath(cal_path)}")
    if not os.path.exists(cal_path):
        logger.error(f"Calibration data file not found: {os.path.abspath(cal_path)}")
        return

    # Load test inputs
    logger.info("Loading test inputs")
    with open(json_path, 'r') as f:
        config = json.load(f)

    sweep_params = config.get("Sweep_Measurement", {}).get("range")
    if not sweep_params:
        logger.error("Sweep_Measurement parameters not found in test_inputs.json")
        return

    start = sweep_params["start_ghz"] * 1e9
    stop = sweep_params["stop_ghz"] * 1e9
    step = sweep_params["step_mhz"] * 1e6
    target_output = sweep_params["power_dbm"]
    tolerance = sweep_params.get("tolerence_db", 0.05)
    expected_gain = sweep_params.get("expected_gain_db", 0.0)
    k575_averages = sweep_params.get("k575_averages", [5, 10, 15, 20, 50])

    freqs = np.arange(start, stop + step, step)
    freqs_ghz = [round(f / 1e9, 3) for f in freqs]
    logger.info(f"Expected frequencies to test (GHz): {freqs_ghz}")

    # Load calibration data
    logger.info("Loading calibration data")
    cal_df = pd.read_excel(cal_path)
    if "Center Frequency (GHz)" not in cal_df.columns:
        logger.error("Calibration data missing 'Center Frequency (GHz)' column")
        return

    # Log available calibration frequencies
    try:
        cal_freqs = cal_df["Center Frequency (GHz)"].astype(float).round(3).unique().tolist()
        logger.info(f"Available calibration frequencies (GHz): {cal_freqs}")
    except ValueError as e:
        logger.error(f"Invalid data in 'Center Frequency (GHz)' column: {str(e)}")
        return

    results = []
    matched_frequencies = set()  # Use set to ensure unique frequencies

    # Initialize instruments
    logger.info("Initializing instruments")
    vsg = VSG()
    vsa = VSA()
    pm = PowerMeter()
    power_servo = PowerServo(vsg, pm, vsa, max_iterations=10, tolerance=tolerance)

    try:
        logger.info("Configuring VSA and VSG initially")
        # Perform initial configuration outside the loop
        initial_freq = freqs[0] if len(freqs) > 0 else start
        initial_freq_ghz = round(initial_freq / 1e9, 3)
        match = cal_df[cal_df["Center Frequency (GHz)"].round(3) == initial_freq_ghz]
        if match.empty:
            logger.error(f"No calibration data for initial frequency {initial_freq_ghz} GHz")
            return
        vsg_offset = float(match["VSG Offset (dB)"].iloc[0])
        vsa_offset = float(match["VSA Offset (dB)"].iloc[0])
        input_offset = float(match["Input Power Offset (dB)"].iloc[0])
        output_offset = float(match["Output Power Offset (dB)"].iloc[0])

        pm.configure(initial_freq, input_offset, output_offset)
        initial_power = target_output - expected_gain
        vsg.configure(initial_freq, initial_power)
        vsa.configure(initial_freq, vsa_offset)

        logger.info(f"Starting sweep over {len(freqs)} frequencies")
        for freq in freqs:
            freq_ghz = round(freq / 1e9, 3)
            if freq_ghz in matched_frequencies:
                logger.warning(f"Frequency {freq_ghz} GHz already processed — skipping")
                continue
            # Match calibration data using exact frequency comparison
            match = cal_df[cal_df["Center Frequency (GHz)"].round(3) == freq_ghz]
            if match.empty:
                logger.warning(f"No calibration data for {freq_ghz} GHz — skipping")
                continue

            matched_frequencies.add(freq_ghz)
            vsg_offset = float(match["VSG Offset (dB)"].iloc[0])
            vsa_offset = float(match["VSA Offset (dB)"].iloc[0])
            input_offset = float(match["Input Power Offset (dB)"].iloc[0])
            output_offset = float(match["Output Power Offset (dB)"].iloc[0])
            logger.info(f"Matched calibration data for {freq_ghz} GHz: VSG Offset={vsg_offset}, VSA Offset={vsa_offset}")

            # Update frequency and power settings
            pm.configure(freq, input_offset, output_offset)
            vsg.vsg.query(f':SOUR1:FREQ:CW {freq}; *OPC?')
            vsa.instr.query(f':SENS:FREQ:CENT {freq}; *OPC?')
            vsa.instr.write(f':DISP:WIND:TRAC:Y:SCAL:RLEV:OFFS {vsa_offset:.2f}')

            # Servo input power to reach target output
            servo_iterations, servo_settle_time = power_servo.servo_power(freq_ghz, target_output, expected_gain)

            # Measure
            corrected_input, corrected_output = pm.measure()
            freq_str = f"{freq:.0f}"
            start_time = time()
            vsa_power, evm_value, evm_time, chan_pow, adj_chan_lower, adj_chan_upper, aclr_time = vsa.measure_evm(freq_str, vsa_offset, corrected_output)

            # Initialize result dictionary
            result = {
                "VSG Setup Time (s)": vsg.setup_time,
                "VSA Setup Time (s)": vsa.setup_time,
                "Center Frequency (GHz)": freq_ghz,
                "Target Output Power (dBm)": target_output,
                "Servo Iterations": servo_iterations,
                "Servo Settle Time (s)": servo_settle_time,
                "Corrected Input Power (dBm)": corrected_input,
                "Corrected Output Power (dBm)": corrected_output,
                "VSA Output Power (dBm)": vsa_power,
                "EVM (dB)": evm_value,
                "EVM Measure Time (s)": evm_time,
                "Channel Power (dBm)": chan_pow,
                "Lower Adjacent ACLR (dB)": adj_chan_lower,
                "Upper Adjacent ACLR (dB)": adj_chan_upper,
                "ACLR Measure Time (s)": aclr_time,
            }

            # Perform K575_EVM and ACLR for each average count
            for avg in k575_averages:
                k575_evm, k575_time, k575_chan_pow, k575_adj_lower, k575_adj_upper, k575_aclr_time = vsa.measure_K575_evm(freq_str, vsa_offset, avg)
                result[f"K575 EVM {avg} avg (dB)"] = k575_evm
                result[f"K575 Time {avg} avg (s)"] = k575_time
                result[f"K575 Channel Power {avg} avg (dBm)"] = k575_chan_pow
                result[f"K575 Lower Adjacent ACLR {avg} avg (dB)"] = k575_adj_lower
                result[f"K575 Upper Adjacent ACLR {avg} avg (dB)"] = k575_adj_upper
                result[f"K575 ACLR Time {avg} avg (s)"] = k575_aclr_time

            elapsed = time() - start_time
            result["Total Elapsed Time (s)"] = round(elapsed, 3)
            results.append(result)
            logger.info(f"Measurement completed at {freq_ghz:.3f} GHz: {result}")

    finally:
        vsg.close()
        #  vsa.close()
        pm.close()
        df = pd.DataFrame(results)
        with pd.ExcelWriter(output_path) as writer:
            df.to_excel(writer, sheet_name='Measurements', index=False)
            if not df.empty:
                stats_rows = [
                    ("Number of Tests", len(df)),
                    ("VSG Setup Time (s) - Mean", df["VSG Setup Time (s)"].mean()),
                    ("VSA Setup Time (s) - Mean", df["VSA Setup Time (s)"].mean()),
                    ("Servo Iterations - Max", df["Servo Iterations"].max()),
                    ("Servo Iterations - Min", df["Servo Iterations"].min()),
                    ("Servo Iterations - Mean", df["Servo Iterations"].mean()),
                    ("EVM (dB) - Max", df["EVM (dB)"].max()),
                    ("EVM (dB) - Min", df["EVM (dB)"].min()),
                    ("EVM (dB) - Mean", df["EVM (dB)"].mean()),
                    ("EVM Measure Time (s) - Max", df["EVM Measure Time (s)"].max()),
                    ("EVM Measure Time (s) - Min", df["EVM Measure Time (s)"].min()),
                    ("EVM Measure Time (s) - Mean", df["EVM Measure Time (s)"].mean()),
                    ("Channel Power (dBm) - Max", df["Channel Power (dBm)"].max()),
                    ("Channel Power (dBm) - Min", df["Channel Power (dBm)"].min()),
                    ("Channel Power (dBm) - Mean", df["Channel Power (dBm)"].mean()),
                    ("Lower Adjacent ACLR (dB) - Max", df["Lower Adjacent ACLR (dB)"].max()),
                    ("Lower Adjacent ACLR (dB) - Min", df["Lower Adjacent ACLR (dB)"].min()),
                    ("Lower Adjacent ACLR (dB) - Mean", df["Lower Adjacent ACLR (dB)"].mean()),
                    ("Upper Adjacent ACLR (dB) - Max", df["Upper Adjacent ACLR (dB)"].max()),
                    ("Upper Adjacent ACLR (dB) - Min", df["Upper Adjacent ACLR (dB)"].min()),
                    ("Upper Adjacent ACLR (dB) - Mean", df["Upper Adjacent ACLR (dB)"].mean()),
                    ("ACLR Measure Time (s) - Max", df["ACLR Measure Time (s)"].max()),
                    ("ACLR Measure Time (s) - Min", df["ACLR Measure Time (s)"].min()),
                    ("ACLR Measure Time (s) - Mean", df["ACLR Measure Time (s)"].mean()),
                ]
                for avg in k575_averages:
                    evm_col = f"K575 EVM {avg} avg (dB)"
                    time_col = f"K575 Time {avg} avg (s)"
                    chan_pow_col = f"K575 Channel Power {avg} avg (dBm)"
                    lower_aclr_col = f"K575 Lower Adjacent ACLR {avg} avg (dB)"
                    upper_aclr_col = f"K575 Upper Adjacent ACLR {avg} avg (dB)"
                    aclr_time_col = f"K575 ACLR Time {avg} avg (s)"
                    stats_rows.extend([
                        (f"{evm_col} - Max", df[evm_col].max()),
                        (f"{evm_col} - Min", df[evm_col].min()),
                        (f"{evm_col} - Mean", df[evm_col].mean()),
                        (f"{time_col} - Max", df[time_col].max()),
                        (f"{time_col} - Min", df[time_col].min()),
                        (f"{time_col} - Mean", df[time_col].mean()),
                        (f"{chan_pow_col} - Max", df[chan_pow_col].max()),
                        (f"{chan_pow_col} - Min", df[chan_pow_col].min()),
                        (f"{chan_pow_col} - Mean", df[chan_pow_col].mean()),
                        (f"{lower_aclr_col} - Max", df[lower_aclr_col].max()),
                        (f"{lower_aclr_col} - Min", df[lower_aclr_col].min()),
                        (f"{lower_aclr_col} - Mean", df[lower_aclr_col].mean()),
                        (f"{upper_aclr_col} - Max", df[upper_aclr_col].max()),
                        (f"{upper_aclr_col} - Min", df[upper_aclr_col].min()),
                        (f"{upper_aclr_col} - Mean", df[upper_aclr_col].mean()),
                        (f"{aclr_time_col} - Max", df[aclr_time_col].max()),
                        (f"{aclr_time_col} - Min", df[aclr_time_col].min()),
                        (f"{aclr_time_col} - Mean", df[aclr_time_col].mean()),
                    ])
                stats_df = pd.DataFrame(stats_rows, columns=["Metric", "Value"])
                stats_df.to_excel(writer, sheet_name='Statistics', index=False)

                # Apply styling after writing
                workbook = writer.book
                worksheet = writer.sheets['Statistics']
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                bold_font = Font(bold=True, size=12)
                for row_idx in range(2, len(stats_rows) + 2):  # Excel rows start at 1, header is row 1
                    metric_cell = worksheet.cell(row=row_idx, column=1)
                    value_cell = worksheet.cell(row=row_idx, column=2)
                    if "Mean" in metric_cell.value:
                        metric_cell.fill = yellow_fill
                        metric_cell.font = bold_font
                        value_cell.fill = yellow_fill
                        value_cell.font = bold_font
            else:
                logger.warning("No data collected; skipping statistics sheet")
        logger.info(f"Saved sweep results to: {os.path.abspath(output_path)}")
        logger.info(f"Processed frequencies (GHz): {sorted(list(matched_frequencies))}")

if __name__ == "__main__":
    run_sweep()